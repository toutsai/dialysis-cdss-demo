from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from src.domain.entities import DialysisSchedule, Patient, Shift


@dataclass(frozen=True)
class ScheduleImportResult:
    patients: list[Patient]
    schedules: list[DialysisSchedule]
    source_updated_at: datetime | None


def load_schedule_workbook(path: str | Path) -> ScheduleImportResult:
    """Load the current Excel-based patient/schedule source.

    TODO(HIS): If a formal patient master or scheduling API becomes available,
    replace this adapter with `hospital/patient_client.py` and
    `hospital/schedule_client.py`.
    """
    wb = load_workbook(path, read_only=False, data_only=True)
    source_updated_at = _extract_source_date(wb.worksheets[1]) if len(wb.worksheets) > 1 else None
    patients_by_chart = _load_patients_from_master_sheet(wb.worksheets[0])
    schedules = _load_schedules_from_roster_sheet(wb.worksheets[1], patients_by_chart, source_updated_at)

    for schedule in schedules:
        if schedule.chart_no not in patients_by_chart:
            patients_by_chart[schedule.chart_no] = Patient(
                chart_no=schedule.chart_no,
                name=schedule.name,
                frequency=schedule.frequency,
                shift=schedule.shift,
                bed=schedule.bed,
                source="schedule_roster",
            )
    return ScheduleImportResult(
        patients=sorted(patients_by_chart.values(), key=lambda p: (p.frequency, p.shift.value, p.bed or "999")),
        schedules=schedules,
        source_updated_at=source_updated_at,
    )


def _load_patients_from_master_sheet(ws) -> dict[str, Patient]:
    out: dict[str, Patient] = {}
    # Sheet "生日重大領藥 " has headers on row 2 in the current workbook:
    # frequency, bed, name, chart_no, birthday...
    for row in ws.iter_rows(min_row=3, values_only=True):
        frequency_raw, bed_raw, name_raw, chart_raw = row[:4]
        name = _clean_text(name_raw)
        chart_no = _normalize_chart_no(chart_raw)
        if not name or not chart_no:
            continue
        frequency = _normalize_frequency(frequency_raw)
        shift = Shift.parse(frequency_raw)
        bed = _clean_bed(bed_raw)
        out[chart_no] = Patient(
            chart_no=chart_no,
            name=name,
            frequency=frequency,
            shift=shift,
            bed=bed,
            source=ws.title,
        )
    return out


def _load_schedules_from_roster_sheet(ws, patients_by_chart: dict[str, Patient], updated_at: datetime | None) -> list[DialysisSchedule]:
    name_to_patient = {p.name: p for p in patients_by_chart.values()}
    schedules: list[DialysisSchedule] = []
    current_beds: dict[int, str] = {}
    current_frequency = ""
    current_shift = Shift.UNKNOWN

    for row in ws.iter_rows(values_only=True):
        marker = _clean_text(row[1] if len(row) > 1 else None)
        if marker and re.search(r"(135|246|一三五|二四六)", marker):
            current_frequency = _normalize_frequency(marker)
            current_shift = Shift.parse(marker)

        # Bed labels are usually numeric values across columns 3..16.
        for idx, value in enumerate(row):
            if idx >= 2 and _looks_like_bed_label(value):
                current_beds[idx] = _clean_bed(value) or ""

        if current_shift == Shift.UNKNOWN:
            continue
        for idx, value in enumerate(row):
            if idx < 2 or not isinstance(value, str) or "\n" not in value:
                continue
            parts = [_clean_text(part) for part in value.splitlines() if _clean_text(part)]
            if not parts:
                continue
            name = parts[0]
            patient = name_to_patient.get(name)
            if patient is None:
                continue
            schedules.append(DialysisSchedule(
                chart_no=patient.chart_no,
                name=name,
                frequency=current_frequency or patient.frequency,
                shift=current_shift,
                bed=current_beds.get(idx) or patient.bed,
                dialyzer=parts[1] if len(parts) > 1 else None,
                dialysate_ca=_extract_ca(parts[2:]),
                source_sheet=ws.title,
                source_updated_at=updated_at,
            ))
    return schedules


def _extract_source_date(ws) -> datetime | None:
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        for value in row:
            if isinstance(value, datetime):
                return value
    return None


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_chart_no(value: object) -> str:
    text = _clean_text(value)
    return text.replace(" ", "")


def _normalize_frequency(value: object) -> str:
    text = _clean_text(value)
    if "246" in text or "二四六" in text:
        return "二四六"
    if "135" in text or "一三五" in text:
        return "一三五"
    if "15" in text or "一五" in text:
        return "一五"
    return text


def _clean_bed(value: object) -> str | None:
    text = _clean_text(value)
    if not text:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return text


def _looks_like_bed_label(value: object) -> bool:
    text = _clean_text(value)
    return bool(re.fullmatch(r"\d+(\.0)?|\d+\s*\([BC]\)(/\([BC]\))?", text))


def _extract_ca(parts: list[str]) -> str | None:
    for part in parts:
        if re.search(r"\b[23]\.5\b|\b3\.0\b|\b2\.5\b", part):
            return part
    return None
